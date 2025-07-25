#!/usr/bin/env python3
"""
Create a backup Excel file with all current passwords and account information.
This creates a copy that you can open safely while the main file is being processed.
"""

import openpyxl
import os
from datetime import datetime
from dotenv import load_dotenv

load_dotenv()

XLSX_PATH = os.getenv("XLSX_PATH", "emails.xlsx")
SHEET_NAME = os.getenv("SHEET_NAME") or None

def create_backup():
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    backup_filename = f"emails_backup_{timestamp}.xlsx"
    
    try:
        # Load the original file
        wb = openpyxl.load_workbook(XLSX_PATH)
        ws = wb[SHEET_NAME] if SHEET_NAME else wb.active
        
        # Create new workbook for backup
        backup_wb = openpyxl.Workbook()
        backup_ws = backup_wb.active
        backup_ws.title = "Email Accounts"
        
        # Add headers
        backup_ws['A1'] = "Email Address"
        backup_ws['B1'] = "Password"
        backup_ws['C1'] = "Status"
        backup_ws['D1'] = "Notes"
        
        # Copy data
        row_count = 0
        for row_num, row in enumerate(ws.iter_rows(min_row=2, min_col=1, max_col=3), 2):
            cell_email, cell_pwd, cell_status = row
            email = (cell_email.value or "").strip()
            password = (cell_pwd.value or "").strip()
            status = (cell_status.value or "").strip()
            
            if email:
                backup_row = row_count + 2
                backup_ws[f'A{backup_row}'] = email
                backup_ws[f'B{backup_row}'] = password
                backup_ws[f'C{backup_row}'] = status
                
                # Add notes for problematic accounts
                if not password and "API Error" in status:
                    backup_ws[f'D{backup_row}'] = "Password needed - account may already exist"
                elif password and status == "OK":
                    backup_ws[f'D{backup_row}'] = "Account created successfully"
                
                row_count += 1
        
        # Format the backup file
        backup_ws.column_dimensions['A'].width = 50
        backup_ws.column_dimensions['B'].width = 20
        backup_ws.column_dimensions['C'].width = 30
        backup_ws.column_dimensions['D'].width = 40
        
        # Save backup
        backup_wb.save(backup_filename)
        
        print(f"✓ Backup created: {backup_filename}")
        print(f"✓ Total accounts: {row_count}")
        
        # Count successful vs failed
        successful = sum(1 for row in ws.iter_rows(min_row=2, min_col=1, max_col=3) 
                        if (row[0].value or "").strip() and (row[1].value or "").strip())
        failed = sum(1 for row in ws.iter_rows(min_row=2, min_col=1, max_col=3) 
                    if (row[0].value or "").strip() and not (row[1].value or "").strip())
        
        print(f"✓ Accounts with passwords: {successful}")
        print(f"⚠ Accounts missing passwords: {failed}")
        
        if failed > 0:
            print(f"\nTo fix missing passwords, close Excel and run:")
            print(f"python recover_passwords.py")
        
    except FileNotFoundError:
        print(f"ERROR: Original file not found: {XLSX_PATH}")
    except Exception as e:
        print(f"ERROR: Failed to create backup: {e}")

if __name__ == "__main__":
    print("Email Account Backup Tool")
    print("=" * 30)
    create_backup()
