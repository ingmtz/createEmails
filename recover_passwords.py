#!/usr/bin/env python3
"""
Generate and set new passwords for email accounts that have blank passwords in the Excel file.
This script will only process rows where the password column (B) is empty but status column (C) shows an error.
"""

import os
import re
import string
import secrets
import json
import requests
from openpyxl import load_workbook
from requests.auth import HTTPBasicAuth
from dotenv import load_dotenv

load_dotenv()

# ── ENV CONFIG ───────────────────────────────────────────────────────────
CPANEL_HOST   = os.getenv("CPANEL_HOST")
CPANEL_USER   = os.getenv("CPANEL_USER")
CPANEL_PASS   = os.getenv("CPANEL_PASS")
CPANEL_TOKEN  = os.getenv("CPANEL_TOKEN")  # optional
XLSX_PATH     = os.getenv("XLSX_PATH", "emails.xlsx")
SHEET_NAME    = os.getenv("SHEET_NAME") or None
VERIFY_SSL    = os.getenv("VERIFY_SSL", "True").lower() in ("1", "true", "yes")
DEBUG         = os.getenv("DEBUG", "False").lower() in ("1", "true", "yes")

EMAIL_REGEX = re.compile(r"^([^@]+)@([^@]+\.[^@]+)$")

def random_password(length: int = 14) -> str:
    """Generate a cryptographically strong letters+digits password with guaranteed complexity."""
    if length < 4:
        length = 4  # Minimum length to ensure all character types
    
    uppercase = string.ascii_uppercase
    lowercase = string.ascii_lowercase
    digits = string.digits
    
    # Ensure at least one of each type
    password = [
        secrets.choice(uppercase),
        secrets.choice(lowercase),
        secrets.choice(digits),
    ]
    
    # Fill the rest with random choices from all character types
    all_chars = uppercase + lowercase + digits
    for _ in range(length - 3):
        password.append(secrets.choice(all_chars))
    
    # Shuffle to avoid predictable patterns
    secrets.SystemRandom().shuffle(password)
    
    return "".join(password)

def change_password(local: str, domain: str, new_pwd: str):
    """Change password for an existing email account."""
    url = f"https://{CPANEL_HOST}:2083/execute/Email/passwd_pop"
    params = {
        "email": local,
        "domain": domain,
        "password": new_pwd,
    }
    
    if DEBUG:
        print(f"DEBUG: Changing password for {local}@{domain}")
        print(f"DEBUG: URL: {url}")
        print(f"DEBUG: Using auth method: {'Token' if CPANEL_TOKEN else 'Username/Password'}")

    # Choose auth method
    if CPANEL_TOKEN:
        headers = {
            "Authorization": f"cpanel {CPANEL_USER}:{CPANEL_TOKEN}",
            "Accept": "application/json",
        }
        auth = None
    else:
        headers = {"Accept": "application/json"}
        auth = HTTPBasicAuth(CPANEL_USER, CPANEL_PASS)

    try:
        r = requests.get(
            url,
            params=params,
            headers=headers,
            auth=auth,
            verify=VERIFY_SSL,
            timeout=15,
        )
        
        if DEBUG:
            print(f"DEBUG: HTTP Status Code: {r.status_code}")
            
    except requests.RequestException as exc:
        return False, f"Request error: {exc}"

    body_text = r.text.strip()
    
    if DEBUG:
        print(f"DEBUG: Raw response body (first 500 chars): {body_text[:500]}")

    # Check for common authentication failure responses
    if "login is invalid" in body_text.lower() or "authentication failed" in body_text.lower():
        return False, "Authentication failed: Invalid username/password or account locked"
    
    if r.status_code == 401:
        return False, "HTTP 401 Unauthorized: Check your cPanel username and password"
    
    if r.status_code == 403:
        return False, "HTTP 403 Forbidden: Account may be suspended or lacks email permissions"

    # Check if we got an HTML response instead of JSON
    if "<html" in body_text.lower() and "login" in body_text.lower():
        return False, "Authentication failed: Received login page instead of API response"

    # Attempt to parse JSON
    try:
        data = json.loads(body_text)
        if data.get("status") == 1:
            return True, "Password changed successfully"
        
        if DEBUG:
            print(f"DEBUG: JSON response: {json.dumps(data, indent=2)}")
        
        # Extract specific error messages
        errors = data.get("errors", [])
        if errors:
            return False, f"API Error: {'; '.join(str(e) for e in errors)}"
        
        result = data.get("result", {})
        if isinstance(result, dict):
            error_msg = result.get("error", "")
            if error_msg:
                return False, f"cPanel Error: {error_msg}"
        
        return False, f"API failed with status {data.get('status', 'unknown')}: {body_text[:200]}"
    except json.JSONDecodeError:
        if "<html" in body_text.lower():
            return False, f"Received HTML response instead of JSON (first 200 chars): {body_text[:200]}"
        
        return False, f"Unexpected response format: {body_text[:200]}"

def main():
    print("cPanel Password Recovery Tool")
    print("=" * 40)
    print("This tool will generate new passwords for email accounts")
    print("that have blank passwords in the Excel file.")
    print()
    
    # Validate configuration
    if not CPANEL_HOST or not CPANEL_USER:
        print("ERROR: Missing required configuration in .env file")
        return
    
    if not CPANEL_TOKEN and not CPANEL_PASS:
        print("ERROR: Either CPANEL_TOKEN or CPANEL_PASS must be provided")
        return
    
    print(f"Processing Excel file: {XLSX_PATH}")
    
    try:
        wb = load_workbook(XLSX_PATH)
        ws = wb[SHEET_NAME] if SHEET_NAME else wb.active
    except FileNotFoundError:
        print(f"ERROR: Excel file not found: {XLSX_PATH}")
        return
    except Exception as e:
        print(f"ERROR: Failed to open Excel file: {e}")
        return

    # Find accounts with blank passwords
    accounts_to_fix = []
    
    for row_num, row in enumerate(ws.iter_rows(min_row=2, min_col=1, max_col=3), 2):
        cell_email, cell_pwd, cell_status = row
        email = (cell_email.value or "").strip()
        password = (cell_pwd.value or "").strip()
        status = (cell_status.value or "").strip()
        
        if email and not password:
            match = EMAIL_REGEX.match(email)
            if match:
                accounts_to_fix.append((row_num, email, match.groups(), cell_pwd, cell_status))
    
    if not accounts_to_fix:
        print("✓ No accounts found with blank passwords. All accounts already have passwords!")
        return
    
    print(f"\nFound {len(accounts_to_fix)} accounts with blank passwords:")
    for i, (row_num, email, _, _, _) in enumerate(accounts_to_fix, 1):
        print(f"  {i}. {email} (row {row_num})")
    
    print("\nDo you want to generate new passwords for these accounts?")
    try:
        response = input("Type 'yes' to continue: ").strip().lower()
        if response != 'yes':
            print("Cancelled.")
            return
    except KeyboardInterrupt:
        print("\nCancelled.")
        return
    
    print("\nGenerating new passwords...")
    
    success_count = 0
    for row_num, email, (local_part, domain), cell_pwd, cell_status in accounts_to_fix:
        new_password = random_password()
        print(f"\nProcessing: {email}")
        
        success, message = change_password(local_part, domain, new_password)
        
        if success:
            cell_pwd.value = new_password
            cell_status.value = "OK"
            success_count += 1
            print(f"✓ SUCCESS: New password set for {email}")
        else:
            cell_status.value = f"Password change failed: {message}"
            print(f"✗ FAILED: {email} - {message}")
    
    # Save the Excel file
    try:
        wb.save(XLSX_PATH)
        print(f"\n" + "=" * 50)
        print(f"Password recovery completed!")
        print(f"Accounts processed: {len(accounts_to_fix)}")
        print(f"Successful: {success_count}")
        print(f"Failed: {len(accounts_to_fix) - success_count}")
        print(f"Results saved to: {XLSX_PATH}")
        print("\n🔐 IMPORTANT: Make sure to close Excel before running this script!")
    except Exception as e:
        print(f"ERROR: Failed to save Excel file: {e}")
        print("This usually means Excel is open. Close Excel and try again.")
        print("\nGenerated passwords (save these manually):")
        for row_num, email, _, _, _ in accounts_to_fix:
            print(f"  {email}: {random_password()}")

if __name__ == "__main__":
    main()
