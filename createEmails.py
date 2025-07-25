#!/usr/bin/env python3
"""
Bulk‑create cPanel e‑mail accounts from an Excel workbook.
Only rows whose column C is *not* "OK" will be processed.
For every attempt, we store either the generated password (column B) and "OK",
or the **entire server response body** (column C) when it is not JSON or reports failure.
Configuration is taken from a `.env` file.
"""

import os
import re
import string
import secrets
import json
import requests
from openpyxl import load_workbook
from requests.auth import HTTPBasicAuth
from dotenv import load_dotenv

load_dotenv()

# ── ENV CONFIG ───────────────────────────────────────────────────────────
CPANEL_HOST   = os.getenv("CPANEL_HOST")
CPANEL_USER   = os.getenv("CPANEL_USER")
CPANEL_PASS   = os.getenv("CPANEL_PASS")
CPANEL_TOKEN  = os.getenv("CPANEL_TOKEN")  # optional
DEFAULT_QUOTA = int(os.getenv("DEFAULT_QUOTA", "0"))
XLSX_PATH     = os.getenv("XLSX_PATH", "bulk_emails.xlsx")
SHEET_NAME    = os.getenv("SHEET_NAME") or None
VERIFY_SSL    = os.getenv("VERIFY_SSL", "True").lower() in ("1", "true", "yes")
DEBUG         = os.getenv("DEBUG", "False").lower() in ("1", "true", "yes")
SKIP_OK_ROWS  = os.getenv("SKIP_OK_ROWS", "True").lower() in ("1", "true", "yes")
SKIP_AUTH_TEST = os.getenv("SKIP_AUTH_TEST", "False").lower() in ("1", "true", "yes")

EMAIL_REGEX = re.compile(r"^([^@]+)@([^@]+\.[^@]+)$")

# ── HELPERS ──────────────────────────────────────────────────────────────

def show_api_token_instructions():
    """Show instructions for creating a cPanel API token."""
    print("\n" + "=" * 60)
    print("HOW TO CREATE A CPANEL API TOKEN")
    print("=" * 60)
    print("1. Log into your cPanel account at:")
    print(f"   https://{CPANEL_HOST}:2083")
    print("")
    print("2. Navigate to: Security > Manage API Tokens")
    print("")
    print("3. Click 'Create' to generate a new token")
    print("")
    print("4. Give it a name like 'Email Script Token'")
    print("")
    print("5. Set expiration (or leave unlimited)")
    print("")
    print("6. Copy the generated token")
    print("")
    print("7. Add this line to your .env file:")
    print("   CPANEL_TOKEN=your_api_token_here")
    print("")
    print("8. Comment out or remove the CPANEL_PASS line")
    print("")
    print("API tokens are more secure than passwords and bypass")
    print("two-factor authentication restrictions for API calls.")
    print("=" * 60)


def random_password(length: int = 14) -> str:
    """Generate a cryptographically strong letters+digits password with guaranteed complexity."""
    if length < 4:
        length = 4  # Minimum length to ensure all character types
    
    uppercase = string.ascii_uppercase
    lowercase = string.ascii_lowercase
    digits = string.digits
    
    # Ensure at least one of each type
    password = [
        secrets.choice(uppercase),
        secrets.choice(lowercase),
        secrets.choice(digits),
    ]
    
    # Fill the rest with random choices from all character types
    all_chars = uppercase + lowercase + digits
    for _ in range(length - 3):
        password.append(secrets.choice(all_chars))
    
    # Shuffle to avoid predictable patterns
    secrets.SystemRandom().shuffle(password)
    
    return "".join(password)


def test_authentication():
    """Test cPanel authentication by calling a simple API function."""
    url = f"https://{CPANEL_HOST}:2083/execute/Email/count_pops"
    
    # Choose auth method
    if CPANEL_TOKEN:
        headers = {
            "Authorization": f"cpanel {CPANEL_USER}:{CPANEL_TOKEN}",
            "Accept": "application/json",
        }
        auth = None
        auth_method = "API Token"
    else:
        headers = {"Accept": "application/json"}
        auth = HTTPBasicAuth(CPANEL_USER, CPANEL_PASS)
        auth_method = "Username/Password"

    print(f"Testing authentication using {auth_method}...")
    print(f"Host: {CPANEL_HOST}")
    print(f"User: {CPANEL_USER}")
    
    try:
        r = requests.get(
            url,
            headers=headers,
            auth=auth,
            verify=VERIFY_SSL,
            timeout=15,
        )
        
        print(f"HTTP Status: {r.status_code}")
        
        # Check if we got an HTML login page (common issue)
        if "html" in r.text.lower() and "login" in r.text.lower():
            print("✗ Received HTML login page instead of JSON API response")
            print("This typically means:")
            print("  - Two-factor authentication is enabled (blocks API access)")
            print("  - API access is disabled for this account")
            print("  - Invalid credentials")
            print("  - Server security settings block API calls")
            print("\nSOLUTIONS:")
            print("1. Create an API Token in cPanel (Security > Manage API Tokens)")
            print("2. Disable two-factor authentication for API calls")
            print("3. Contact your hosting provider to enable API access")
            return False
        
        if r.status_code == 200:
            try:
                data = json.loads(r.text)
                if data.get("status") == 1:
                    count = data.get("result", {}).get("data", "unknown")
                    print(f"✓ Authentication successful! Found {count} email accounts.")
                    return True
                else:
                    print(f"✗ API call failed: {data}")
                    return False
            except json.JSONDecodeError:
                print(f"✗ Non-JSON response: {r.text[:200]}")
                return False
        else:
            print(f"✗ HTTP Error {r.status_code}: {r.text[:200]}")
            return False
            
    except requests.RequestException as exc:
        print(f"✗ Connection error: {exc}")
        return False


def create_mailbox(local: str, domain: str, pwd: str, quota_mb: int = 0):
    """Try to create the mailbox and return (success_bool, message_text)."""
    url = f"https://{CPANEL_HOST}:2083/execute/Email/add_pop"
    params = {
        "email":    local,
        "domain":   domain,
        "password": pwd,
        "quota":    quota_mb,
    }
    
    if DEBUG:
        print(f"DEBUG: Attempting to create {local}@{domain}")
        print(f"DEBUG: URL: {url}")
        print(f"DEBUG: Using auth method: {'Token' if CPANEL_TOKEN else 'Username/Password'}")

    # Choose auth method
    if CPANEL_TOKEN:
        headers = {
            "Authorization": f"cpanel {CPANEL_USER}:{CPANEL_TOKEN}",
            "Accept": "application/json",
        }
        auth = None
    else:
        headers = {"Accept": "application/json"}
        auth = HTTPBasicAuth(CPANEL_USER, CPANEL_PASS)

    try:
        r = requests.get(
            url,
            params=params,
            headers=headers,
            auth=auth,
            verify=VERIFY_SSL,
            timeout=15,
        )
        
        if DEBUG:
            print(f"DEBUG: HTTP Status Code: {r.status_code}")
            print(f"DEBUG: Response Headers: {dict(r.headers)}")
            
    except requests.RequestException as exc:
        return False, f"Request error: {exc}"

    body_text = r.text.strip()
    
    if DEBUG:
        print(f"DEBUG: Raw response body (first 500 chars): {body_text[:500]}")

    # Check for common authentication failure HTML responses
    if "login is invalid" in body_text.lower() or "authentication failed" in body_text.lower():
        return False, "Authentication failed: Invalid username/password or account locked"
    
    if r.status_code == 401:
        return False, "HTTP 401 Unauthorized: Check your cPanel username and password"
    
    if r.status_code == 403:
        return False, "HTTP 403 Forbidden: Account may be suspended or lacks email permissions"

    # Check if we got an HTML response instead of JSON (but be less strict than the test function)
    if "<html" in body_text.lower() and "login" in body_text.lower():
        return False, "Authentication failed: Received login page instead of API response"

    # Attempt to parse JSON. If that fails or JSON reports failure, we keep the raw body.
    try:
        data = json.loads(body_text)
        if data.get("status") == 1:
            return True, "OK"
        # If JSON but failed, provide more specific error information
        if DEBUG:
            print(f"DEBUG: JSON response: {json.dumps(data, indent=2)}")
        
        # Extract specific error messages from the JSON response
        errors = data.get("errors", [])
        if errors:
            return False, f"API Error: {'; '.join(str(e) for e in errors)}"
        
        # Check for specific error conditions
        result = data.get("result", {})
        if isinstance(result, dict):
            error_msg = result.get("error", "")
            if error_msg:
                return False, f"cPanel Error: {error_msg}"
        
        return False, f"API failed with status {data.get('status', 'unknown')}: {body_text[:200]}"
    except json.JSONDecodeError:
        # Not JSON - check if it's HTML or other format
        if "<html" in body_text.lower():
            return False, f"Received HTML response instead of JSON (first 200 chars): {body_text[:200]}"
        
        # If it's not HTML and not JSON, it might be a plain text error
        return False, f"Unexpected response format: {body_text[:200]}"


# ── MAIN LOOP ────────────────────────────────────────────────────────────

def main() -> None:
    print("cPanel Email Account Creator")
    print("=" * 40)
    
    # Validate configuration
    if not CPANEL_HOST or not CPANEL_USER:
        print("ERROR: Missing required configuration in .env file")
        print("Required: CPANEL_HOST, CPANEL_USER")
        return
    
    if not CPANEL_TOKEN and not CPANEL_PASS:
        print("ERROR: Either CPANEL_TOKEN or CPANEL_PASS must be provided in .env file")
        return
    
    # Test authentication before proceeding (unless skipped)
    if not SKIP_AUTH_TEST:
        auth_test_passed = test_authentication()
        if not auth_test_passed:
            print("\nAuthentication test failed, but this might be due to endpoint restrictions.")
            print("Do you want to proceed anyway? The email creation might still work.")
            
            # Show API token instructions if using password authentication
            if not CPANEL_TOKEN:
                show_api_token_instructions()
            
            # Ask user if they want to continue
            try:
                response = input("\nDo you want to continue anyway? (y/N): ").strip().lower()
                if response not in ['y', 'yes']:
                    print("Exiting...")
                    return
            except KeyboardInterrupt:
                print("\nExiting...")
                return
            
            print("\nProceeding with email creation despite authentication test failure...")
    else:
        print("Skipping authentication test as requested...")
    
    print(f"\nProcessing Excel file: {XLSX_PATH}")
    
    try:
        wb = load_workbook(XLSX_PATH)
        ws = wb[SHEET_NAME] if SHEET_NAME else wb.active
    except FileNotFoundError:
        print(f"ERROR: Excel file not found: {XLSX_PATH}")
        return
    except Exception as e:
        print(f"ERROR: Failed to open Excel file: {e}")
        return

    processed_count = 0
    success_count = 0
    
    for row in ws.iter_rows(min_row=2, min_col=1, max_col=3):
        cell_email, cell_pwd, cell_status = row
        email = (cell_email.value or "").strip()

        if not email:
            cell_status.value = "Skipped: empty cell"
            continue

        # Skip rows already marked OK if flag enabled
        if SKIP_OK_ROWS and (cell_status.value or "").strip().upper() == "OK":
            if DEBUG:
                print(f"{email}: skipped (already OK)")
            continue

        match = EMAIL_REGEX.match(email)
        if not match:
            cell_status.value = "Invalid email format"
            continue

        local_part, domain = match.groups()
        password = random_password()

        processed_count += 1
        print(f"\nProcessing ({processed_count}): {email}")
        
        ok, msg = create_mailbox(local_part, domain, password, DEFAULT_QUOTA)
        if ok:
            cell_pwd.value = password
            success_count += 1
            print(f"✓ SUCCESS: {email}")
        else:
            print(f"✗ FAILED: {email} - {msg[:100]}{'…' if len(msg) > 100 else ''}")
        
        cell_status.value = msg

    try:
        wb.save(XLSX_PATH)
        print(f"\n" + "=" * 50)
        print(f"Processing completed!")
        print(f"Total processed: {processed_count}")
        print(f"Successful: {success_count}")
        print(f"Failed: {processed_count - success_count}")
        print(f"Results saved to: {XLSX_PATH}")
    except Exception as e:
        print(f"ERROR: Failed to save Excel file: {e}")
        print("Results were processed but not saved!")


if __name__ == "__main__":
    main()
